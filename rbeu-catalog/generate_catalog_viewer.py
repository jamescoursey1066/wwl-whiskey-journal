#!/usr/bin/env python3
"""
Build an interactive HTML viewer for the xRBEU Catalog.
Reads RBEU Catalog.xlsx, computes xRBEU using the same exponential formula
as add_whiskey.py, and emits catalog_viewer.html with filter/sort/search.
"""

import os
import json
import math
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SPREADSHEET = os.path.join(HERE, "RBEU Catalog.xlsx")
OUTPUT_HTML = os.path.join(HERE, "catalog_viewer.html")
IMAGES_JSON = os.path.join(HERE, "bottle_images.json")

WT_ROW = 30
TOP_RATING = 100.0
TOP_PRICE = 250.00


def tier(v):
    if v >= 1.75: return "S"
    if v >= 1.00: return "A"
    if v >= 0.85: return "B"
    if v >= 0.65: return "C"
    if v >= 0.50: return "D"
    return "F"


def main():
    images = {}
    if os.path.exists(IMAGES_JSON):
        with open(IMAGES_JSON, encoding="utf-8") as f:
            images = json.load(f)

    wb = load_workbook(SPREADSHEET, data_only=False)
    ws = wb.active

    wt_rating = float(ws.cell(row=WT_ROW, column=4).value)
    wt_price = float(ws.cell(row=WT_ROW, column=5).value)
    k = math.log(TOP_PRICE / wt_price) / (TOP_RATING - wt_rating)

    rows = []
    for row_num in range(5, ws.max_row + 1):
        name = ws.cell(row=row_num, column=3).value
        rating = ws.cell(row=row_num, column=4).value
        price = ws.cell(row=row_num, column=5).value
        bbb = ws.cell(row=row_num, column=8).value or ""
        notes = ws.cell(row=row_num, column=9).value or ""
        if name is None or rating is None or price is None:
            continue
        rating = float(rating)
        price = float(price)
        fair = wt_price * math.exp(k * (rating - wt_rating))
        rbeu = fair / price
        clean_name = str(name).strip()
        rows.append({
            "name": clean_name,
            "rating": rating,
            "price": price,
            "fair": round(fair, 2),
            "rbeu": round(rbeu, 4),
            "tier": tier(rbeu),
            "bbb": str(bbb).strip(),
            "notes": str(notes).strip(),
            "row": row_num,
            "img": images.get(clean_name) or None,
        })

    rows.sort(key=lambda x: -x["rbeu"])
    for i, r in enumerate(rows, 1):
        r["rank"] = i

    payload = {
        "anchors": {
            "wt_rating": wt_rating,
            "wt_price": wt_price,
            "top_rating": TOP_RATING,
            "top_price": TOP_PRICE,
            "k": round(k, 6),
        },
        "rows": rows,
    }

    html = HTML_TEMPLATE.replace("__DATA__", json.dumps(payload))
    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print("Wrote {} ({} whiskeys)".format(OUTPUT_HTML, len(rows)))


HTML_TEMPLATE = (
    "<!DOCTYPE html>\n"
    "<html lang=\"en\">\n"
    "<head>\n"
    "<meta charset=\"UTF-8\">\n"
    "<title>xRBEU Catalog Viewer</title>\n"
    "<style>\n"
    "  :root {\n"
    "    --bg: #fafaf7; --panel: #ffffff; --ink: #1a1a1a; --muted: #6b6b6b;\n"
    "    --line: #e5e3dd; --accent: #6b3e1f;\n"
    "    --tier-S: #d4af37; --tier-A: #4a7c4a; --tier-B: #5b8def;\n"
    "    --tier-C: #c9a36b; --tier-D: #a86a4a; --tier-F: #9b3838;\n"
    "  }\n"
    "  * { box-sizing: border-box; }\n"
    "  body { margin: 0; font-family: -apple-system, BlinkMacSystemFont, \"Segoe UI\", system-ui, sans-serif;\n"
    "    background: var(--bg); color: var(--ink); line-height: 1.4; font-size: 14px; }\n"
    "  header { background: var(--panel); border-bottom: 1px solid var(--line);\n"
    "    padding: 16px 24px; position: sticky; top: 0; z-index: 10; }\n"
    "  h1 { margin: 0 0 4px 0; font-size: 20px; }\n"
    "  .subtitle { color: var(--muted); font-size: 12px; margin-bottom: 12px; }\n"
    "  .controls { display: flex; gap: 12px; flex-wrap: wrap; align-items: center; }\n"
    "  .control { display: flex; flex-direction: column; gap: 4px; }\n"
    "  .control label { font-size: 11px; color: var(--muted); text-transform: uppercase; letter-spacing: 0.5px; }\n"
    "  input[type=\"text\"] { padding: 6px 10px; border: 1px solid var(--line); border-radius: 4px;\n"
    "    background: var(--bg); font-size: 14px; min-width: 160px; }\n"
    "  .chip-row { display: flex; gap: 6px; flex-wrap: wrap; }\n"
    "  .chip { padding: 4px 10px; border: 1px solid var(--line); border-radius: 16px;\n"
    "    cursor: pointer; user-select: none; font-size: 12px; background: var(--panel); transition: all 0.15s; }\n"
    "  .chip.active { background: var(--accent); color: white; border-color: var(--accent); }\n"
    "  .stats { display: flex; gap: 20px; padding: 12px 24px; background: var(--panel);\n"
    "    border-bottom: 1px solid var(--line); font-size: 12px; color: var(--muted); flex-wrap: wrap; }\n"
    "  .stat strong { color: var(--ink); font-size: 14px; }\n"
    "  table { width: 100%; border-collapse: collapse; background: var(--panel); }\n"
    "  th, td { padding: 8px 12px; text-align: left; border-bottom: 1px solid var(--line); vertical-align: middle; }\n"
    "  th { background: #f4f2ec; font-weight: 600; font-size: 12px; text-transform: uppercase;\n"
    "    letter-spacing: 0.5px; cursor: pointer; position: sticky; top: 0; z-index: 1; user-select: none; }\n"
    "  th:hover { background: #ebe8df; }\n"
    "  th.img-th { width: 52px; min-width: 52px; cursor: default; padding: 4px 8px; text-align: center; }\n"
    "  th.img-th:hover { background: #f4f2ec; }\n"
    "  th.sorted { color: var(--accent); }\n"
    "  th.sorted::after { content: \"\"; }\n"
    "  th.sorted.asc::after { content: \" \u25b2\"; font-size: 10px; }\n"
    "  th.sorted.desc::after { content: \" \u25bc\"; font-size: 10px; }\n"
    "  td.num, th.num { text-align: right; font-variant-numeric: tabular-nums; }\n"
    "  td.center, th.center { text-align: center; }\n"
    "  tbody tr:hover { background: #faf7ee; }\n"
    "  .tier-badge { display: inline-block; min-width: 22px; padding: 2px 8px; border-radius: 4px;\n"
    "    color: white; font-weight: 700; text-align: center; font-size: 12px; }\n"
    "  .tier-S { background: var(--tier-S); color: #3a2d00; }\n"
    "  .tier-A { background: var(--tier-A); }\n"
    "  .tier-B { background: var(--tier-B); }\n"
    "  .tier-C { background: var(--tier-C); }\n"
    "  .tier-D { background: var(--tier-D); }\n"
    "  .tier-F { background: var(--tier-F); }\n"
    "  .bbb { font-size: 11px; padding: 2px 6px; border-radius: 3px; background: #f0ede5; color: var(--muted); }\n"
    "  .bbb-Bottle { background: #d9efd9; color: #1f5c1f; }\n"
    "  .bbb-Bar { background: #fff2cc; color: #6b5208; }\n"
    "  .bbb-Bust { background: #f9d4d4; color: #7c1e1e; }\n"
    "  .name { font-weight: 500; }\n"
    "  .name .notes { color: var(--muted); font-size: 11px; font-style: italic; margin-top: 2px; }\n"
    "  .empty { padding: 32px; text-align: center; color: var(--muted); }\n"
    "  footer { padding: 16px 24px; color: var(--muted); font-size: 11px; text-align: center;\n"
    "    border-top: 1px solid var(--line); }\n"
    "  td.img-td { width: 52px; min-width: 52px; padding: 4px 8px; text-align: center; }\n"
    "  .thumb { width: 36px; height: 54px; object-fit: contain; cursor: zoom-in;\n"
    "    display: block; margin: 0 auto; border-radius: 3px; }\n"
    "  .no-img { display: block; width: 36px; height: 54px; margin: 0 auto;\n"
    "    background: url(\"data:image/svg+xml,%3Csvg%20xmlns%3D%22http%3A%2F%2Fwww.w3.org%2F2000%2Fsvg%22%20viewBox%3D%220%200%2036%2054%22%3E%3Cg%20fill%3D%22%23d4d0c8%22%20stroke%3D%22%23b8b4ac%22%20stroke-width%3D%221.5%22%3E%3Crect%20x%3D%2213%22%20y%3D%221%22%20width%3D%2210%22%20height%3D%225%22%20rx%3D%221.5%22%2F%3E%3Crect%20x%3D%2212%22%20y%3D%226%22%20width%3D%2212%22%20height%3D%228%22%2F%3E%3Cpath%20d%3D%22M12%2014C4%2020%204%2026%204%2030v18c0%204%204%205%2014%205s14-1%2014-5V30c0-4%200-10-8-16z%22%2F%3E%3C%2Fg%3E%3C%2Fsvg%3E\")"
    "    center/contain no-repeat; }\n"
    "  #modal { display: none; position: fixed; inset: 0; z-index: 100;\n"
    "    background: rgba(0,0,0,0.78); align-items: center; justify-content: center; cursor: zoom-out; }\n"
    "  #modal img { max-width: 88vw; max-height: 88vh; object-fit: contain;\n"
    "    border-radius: 6px; box-shadow: 0 8px 40px rgba(0,0,0,0.5); cursor: default; }\n"
    "  #modal-close { position: absolute; top: 14px; right: 22px; color: white; font-size: 30px;\n"
    "    cursor: pointer; line-height: 1; background: none; border: none; padding: 0; }\n"
    "</style>\n"
    "</head>\n"
    "<body>\n"
    "<header>\n"
    "  <h1>xRBEU Catalog</h1>\n"
    "  <div class=\"subtitle\" id=\"anchor-info\">Loading\u2026</div>\n"
    "  <div class=\"controls\">\n"
    "    <div class=\"control\"><label>Search</label>\n"
    "      <input type=\"text\" id=\"search\" placeholder=\"Whiskey name\u2026\" /></div>\n"
    "    <div class=\"control\"><label>Tier</label><div class=\"chip-row\" id=\"tier-chips\"></div></div>\n"
    "    <div class=\"control\"><label>BBB</label><div class=\"chip-row\" id=\"bbb-chips\"></div></div>\n"
    "    <div class=\"control\"><label>&nbsp;</label>\n"
    "      <button id=\"reset\" style=\"padding:6px 12px;cursor:pointer;border:1px solid var(--line);background:var(--bg);border-radius:4px;\">Reset filters</button></div>\n"
    "  </div>\n"
    "</header>\n"
    "<div class=\"stats\" id=\"stats\"></div>\n"
    "<table>\n"
    "  <colgroup><col style=\"width:52px\"><col><col><col><col><col><col><col><col></colgroup>\n"
    "  <thead><tr id=\"header-row\"></tr></thead>\n"
    "  <tbody id=\"tbody\"></tbody>\n"
    "</table>\n"
    "<footer>Generated from xRBEU Catalog.xlsx \u2014 re-run <code>generate_catalog_viewer.py</code> after editing.</footer>\n"
    "<div id=\"modal\">\n"
    "  <button id=\"modal-close\" onclick=\"document.getElementById('modal').style.display='none'\">&#x2715;</button>\n"
    "  <img id=\"modal-img\" src=\"\" alt=\"\" onclick=\"event.stopPropagation()\">\n"
    "</div>\n"
    "<script>\n"
    "const DATA = __DATA__;\n"
    "const COLUMNS = [\n"
    "  { key: \"rank\",   label: \"Rank\",   cls: \"center\",   fmt: function(v) { return v; } },\n"
    "  { key: \"tier\",   label: \"Tier\",   cls: \"center\",   fmt: function(v) { return '<span class=\"tier-badge tier-' + v + '\">' + v + '</span>'; } },\n"
    "  { key: \"name\",   label: \"Whiskey\",cls: \"\",         fmt: function(v, row) { return '<div class=\"name\">' + escapeHtml(v) + (row.notes ? '<div class=\"notes\">' + escapeHtml(row.notes) + '</div>' : '') + '</div>'; } },\n"
    "  { key: \"rating\", label: \"Rating\", cls: \"num\",      fmt: function(v) { return String(Math.round(v)); } },\n"
    "  { key: \"price\",  label: \"Price\",  cls: \"num\",      fmt: function(v) { return '$' + v.toFixed(2); } },\n"
    "  { key: \"fair\",   label: \"Fair $\", cls: \"num\",      fmt: function(v) { return '$' + v.toFixed(2); } },\n"
    "  { key: \"rbeu\",   label: \"xRBEU\",  cls: \"num\",      fmt: function(v) { return '<strong>' + v.toFixed(2) + '</strong>'; } },\n"
    "  { key: \"bbb\",    label: \"BBB\",    cls: \"center\",   fmt: function(v) { return v ? '<span class=\"bbb bbb-' + v + '\">' + v + '</span>' : '\u2014'; } },\n"
    "];\n"
    "const TIERS = [\"S\",\"A\",\"B\",\"C\",\"D\",\"F\"];\n"
    "const BBB_OPTS = [\"Bottle\",\"Bar\",\"Bust\"];\n"
    "const STORAGE_KEY = \"rbeu_catalog_filters_v1\";\n"
    "const stored = (function() { try { return JSON.parse(localStorage.getItem(STORAGE_KEY)) || {}; } catch(e) { return {}; } })();\n"
    "const state = { search: stored.search || \"\", tiers: new Set(stored.tiers || []),\n"
    "  bbbs: new Set(stored.bbbs || []), sortKey: stored.sortKey || \"rank\", sortDir: stored.sortDir || \"asc\" };\n"
    "function persist() {\n"
    "  localStorage.setItem(STORAGE_KEY, JSON.stringify({ search: state.search,\n"
    "    tiers: Array.from(state.tiers), bbbs: Array.from(state.bbbs),\n"
    "    sortKey: state.sortKey, sortDir: state.sortDir })); }\n"
    "function escapeHtml(s) {\n"
    "  return String(s).replace(/[&<>\"']/g, function(c) {\n"
    "    return {\"&\":\"&amp;\",\"<\":\"&lt;\",\">\":\"&gt;\",'\"':'&quot;',\"'\": \"&#39;\"}[c]; }); }\n"
    "function imgCell(r) {\n"
    "  if (r.img) {\n"
    "    var isShopify = r.img.indexOf('cdn/shop/') !== -1;\n"
    "    var src = isShopify ? r.img + '&width=80' : r.img;\n"
    "    var full = isShopify ? r.img + '&width=600' : r.img;\n"
    "    return '<img src=\"' + src + '\" loading=\"lazy\" decoding=\"async\" class=\"thumb\" data-full=\"' + full + '\" alt=\"\">'; }\n"
    "  return '<span class=\"no-img\"></span>'; }\n"
    "function renderHeader() {\n"
    "  const row = document.getElementById(\"header-row\");\n"
    "  let html = '<th class=\"img-th\"></th>';\n"
    "  html += COLUMNS.map(function(c) {\n"
    "    const sorted = state.sortKey === c.key;\n"
    "    const cls = (c.cls || \"\") + (sorted ? \" sorted \" + state.sortDir : \"\");\n"
    "    return '<th data-key=\"' + c.key + '\" class=\"' + cls.trim() + '\">' + c.label + '</th>'; }).join(\"\");\n"
    "  row.innerHTML = html;\n"
    "  row.querySelectorAll(\"th[data-key]\").forEach(function(th) {\n"
    "    th.addEventListener(\"click\", function() {\n"
    "      const key = th.dataset.key;\n"
    "      if (state.sortKey === key) { state.sortDir = state.sortDir === \"asc\" ? \"desc\" : \"asc\"; }\n"
    "      else { state.sortKey = key; state.sortDir = (key===\"name\"||key===\"tier\"||key===\"bbb\") ? \"asc\" : \"desc\"; }\n"
    "      persist(); render(); }); }); }\n"
    "function renderChips() {\n"
    "  const tierEl = document.getElementById(\"tier-chips\");\n"
    "  tierEl.innerHTML = TIERS.map(function(t) {\n"
    "    return '<span class=\"chip ' + (state.tiers.has(t) ? \"active\" : \"\") + '\" data-tier=\"' + t + '\">' + t + '</span>'; }).join(\"\");\n"
    "  tierEl.querySelectorAll(\".chip\").forEach(function(el) {\n"
    "    el.addEventListener(\"click\", function() {\n"
    "      const t = el.dataset.tier; if (state.tiers.has(t)) { state.tiers.delete(t); } else { state.tiers.add(t); }\n"
    "      persist(); render(); }); });\n"
    "  const bbbEl = document.getElementById(\"bbb-chips\");\n"
    "  bbbEl.innerHTML = BBB_OPTS.map(function(b) {\n"
    "    return '<span class=\"chip ' + (state.bbbs.has(b) ? \"active\" : \"\") + '\" data-bbb=\"' + b + '\">' + b + '</span>'; }).join(\"\");\n"
    "  bbbEl.querySelectorAll(\".chip\").forEach(function(el) {\n"
    "    el.addEventListener(\"click\", function() {\n"
    "      const b = el.dataset.bbb; if (state.bbbs.has(b)) { state.bbbs.delete(b); } else { state.bbbs.add(b); }\n"
    "      persist(); render(); }); }); }\n"
    "function filtered() {\n"
    "  const q = state.search.trim().toLowerCase();\n"
    "  return DATA.rows.filter(function(r) {\n"
    "    if (state.tiers.size > 0 && !state.tiers.has(r.tier)) return false;\n"
    "    if (state.bbbs.size > 0 && !state.bbbs.has(r.bbb)) return false;\n"
    "    if (q && !r.name.toLowerCase().includes(q) && !(r.notes||\"\").toLowerCase().includes(q)) return false;\n"
    "    return true; }); }\n"
    "function sortedRows(rows) {\n"
    "  const k = state.sortKey; const d = state.sortDir === \"asc\" ? 1 : -1;\n"
    "  const tierRank = {S:0,A:1,B:2,C:3,D:4,F:5};\n"
    "  return rows.slice().sort(function(a,b) {\n"
    "    let av = a[k], bv = b[k];\n"
    "    if (k === \"tier\") { av = tierRank[av]; bv = tierRank[bv]; }\n"
    "    if (typeof av === \"string\") return av.localeCompare(bv) * d;\n"
    "    return ((av||0)-(bv||0))*d; }); }\n"
    "function renderStats(visible) {\n"
    "  const el = document.getElementById(\"stats\");\n"
    "  if (visible.length === 0) { el.innerHTML = '<div class=\"stat\">No matches</div>'; return; }\n"
    "  const counts = TIERS.map(function(t) { return {t:t, n:visible.filter(function(r){return r.tier===t;}).length}; });\n"
    "  const vals = visible.map(function(r){return r.rbeu;}).sort(function(a,b){return a-b;});\n"
    "  const median = vals.length%2 ? vals[(vals.length-1)/2] : (vals[vals.length/2-1]+vals[vals.length/2])/2;\n"
    "  const avgPrice = visible.reduce(function(s,r){return s+r.price;},0)/visible.length;\n"
    "  el.innerHTML = '<div class=\"stat\"><strong>' + visible.length + '</strong> shown / ' + DATA.rows.length + ' total</div>' +\n"
    "    '<div class=\"stat\">Median xRBEU: <strong>' + median.toFixed(2) + '</strong></div>' +\n"
    "    '<div class=\"stat\">Avg price: <strong>$' + avgPrice.toFixed(0) + '</strong></div>' +\n"
    "    counts.map(function(c){return '<div class=\"stat\">'+c.t+': <strong>'+c.n+'</strong></div>';}).join(\"\"); }\n"
    "function renderTable(visible) {\n"
    "  const tbody = document.getElementById(\"tbody\");\n"
    "  if (visible.length === 0) {\n"
    "    tbody.innerHTML = '<tr><td colspan=\"' + (COLUMNS.length+1) + '\" class=\"empty\">No whiskeys match your filters.</td></tr>'; return; }\n"
    "  tbody.innerHTML = visible.map(function(r) {\n"
    "    return '<tr><td class=\"img-td\">' + imgCell(r) + '</td>' +\n"
    "      COLUMNS.map(function(c) {\n"
    "        return '<td class=\"' + (c.cls||\"\") + '\">' + c.fmt(r[c.key],r) + '</td>'; }).join(\"\") + '</tr>'; }).join(\"\"); }\n"
    "function render() { renderHeader(); renderChips(); const visible=sortedRows(filtered()); renderStats(visible); renderTable(visible); }\n"
    "document.getElementById(\"tbody\").addEventListener(\"click\", function(e) {\n"
    "  const img = e.target.closest(\".thumb\"); if (!img) return;\n"
    "  document.getElementById(\"modal-img\").src = img.getAttribute(\"data-full\");\n"
    "  document.getElementById(\"modal\").style.display = \"flex\"; });\n"
    "document.getElementById(\"modal\").addEventListener(\"click\", function() {\n"
    "  document.getElementById(\"modal\").style.display = \"none\"; });\n"
    "document.getElementById(\"search\").value = state.search;\n"
    "document.getElementById(\"search\").addEventListener(\"input\", function(e) { state.search=e.target.value; persist(); render(); });\n"
    "document.getElementById(\"reset\").addEventListener(\"click\", function() {\n"
    "  state.search=\"\"; state.tiers.clear(); state.bbbs.clear();\n"
    "  state.sortKey=\"rank\"; state.sortDir=\"asc\";\n"
    "  document.getElementById(\"search\").value=\"\"; persist(); render(); });\n"
    "document.getElementById(\"anchor-info\").textContent =\n"
    "  \"Anchors: WT R\" + DATA.anchors.wt_rating + \" = $\" + DATA.anchors.wt_price.toFixed(2) + \" \u00b7 \" +\n"
    "  \"Top R\" + DATA.anchors.top_rating + \" = $\" + DATA.anchors.top_price.toFixed(2) + \" \u00b7 \" +\n"
    "  \"k = \" + DATA.anchors.k;\n"
    "render();\n"
    "</script>\n"
    "</body>\n"
    "</html>\n"
)

if __name__ == "__main__":
    main()
