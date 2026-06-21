"""
RBEU Catalog - Monthly Market Price Updater
Updates Market Price (col J) for ALL 89 expressions based on current
Madison / Milwaukee / Chicagoland retail market data.
Street RBEU (col K) and Street Tier (col L) recalculate automatically via Excel formula.

Usage: python update_market_prices.py
"""

import re
from openpyxl import load_workbook

CATALOG_PATH = "/sessions/vigilant-blissful-cannon/mnt/RBEU Catalog/RBEU Catalog.xlsx"

# High-volatility expressions: actively search for current local prices each month.
# Tuple: (partial name match, search query, fallback price)
HIGH_VOLATILITY = [
    ("buffalo trace bourbon",
     "Buffalo Trace bourbon retail price Chicago Illinois Wisconsin {year}",
     32.00),
    ("colonel e.h. taylor small batch",
     "Colonel EH Taylor Small Batch retail price Chicago Wisconsin {year}",
     90.00),
    ("elmer t. lee single barrel",
     "Elmer T Lee Single Barrel retail price Chicago Illinois {year}",
     90.00),
    ("hibiki japanese harmony",
     "Hibiki Japanese Harmony retail price Chicago {year}",
     95.00),
    ("nikka yoichi 10 year single malt",
     "Nikka Yoichi 10 Year retail price Chicago Illinois {year}",
     175.00),
    ("ichiro malt & grain japanese whiskey",
     "Ichiro Malt Grain whiskey retail price Chicago {year}",
     115.00),
    ("westland garryana 9th edition",
     "Westland Garryana retail price Chicago {year}",
     150.00),
    ("redbreast 12 year old irish whiskey",
     "Redbreast 12 year retail price Chicago Wisconsin {year}",
     65.00),
    ("knob creek small batch 9 year",
     "Knob Creek 9 year retail price Chicago Wisconsin {year}",
     45.00),
    ("kujira kyojin okinawa japanese whisky",
     "Kujira Kyojin retail price Chicago {year}",
     75.00),
    ("ohishi 8 year ex",
     "Ohishi 8 year sherry cask retail price USA {year}",
     79.99),
    ("indri trini three wood",
     "Indri Trini Three Wood retail price USA {year}",
     69.99),
    ("milk & honey single cask single malt",
     "Milk Honey Single Cask retail price USA {year}",
     90.00),
    ("orphan barrel rhetoric 22 year",
     "Orphan Barrel Rhetoric 22 Year retail price {year}",
     140.00),
    ("calumet farm 16 year",
     "Calumet Farm 16 Year retail price {year}",
     159.99),
]


def extract_price(text: str) -> float | None:
    """Extract a plausible 750ml bottle price from search result text."""
    matches = re.findall(r'\$(\d{1,3}(?:\.\d{2})?)', text)
    prices = [float(m) for m in matches if 10 < float(m) < 500]
    if not prices:
        return None
    prices.sort()
    return prices[len(prices) // 2]  # median to avoid outlier skew


def search_price(query: str, fallback: float, year: int) -> float:
    """
    During a scheduled Claude run, Claude substitutes real WebSearch results here.
    In standalone use, returns the fallback price.
    """
    print(f"  [search] {query.format(year=year)}")
    print(f"           → using fallback ${fallback:.2f}")
    return fallback


def build_market_prices(searched_prices: dict[str, float], catalog_rows: list) -> dict[int, float]:
    """
    For every catalog row, assign a market price:
    - high-volatility expressions: use searched price
    - all others: use catalog MSRP
    Returns {row_number: market_price}
    """
    result = {}
    for row_num, whiskey, msrp in catalog_rows:
        w = whiskey.strip().lower()
        market_price = msrp  # default
        for key, price in searched_prices.items():
            if key in w:
                market_price = price
                break
        result[row_num] = market_price
    return result


def run():
    import datetime
    year = datetime.date.today().year
    print(f"RBEU Catalog — Monthly Market Price Update ({year})")
    print("Markets: Madison WI / Milwaukee WI / Chicagoland IL\n")

    # Step 1: search for high-volatility prices
    searched = {}
    for name_key, query, fallback in HIGH_VOLATILITY:
        price = search_price(query, fallback, year)
        searched[name_key] = price

    # Step 2: load catalog and set all market prices
    wb = load_workbook(CATALOG_PATH)  # NOT data_only — preserves col K formulas
    ws = wb.active

    catalog_rows = []
    for row_num in range(5, ws.max_row + 1):
        whiskey = ws.cell(row=row_num, column=3).value
        msrp = ws.cell(row=row_num, column=5).value
        if whiskey and msrp:
            catalog_rows.append((row_num, whiskey, msrp))

    market_prices = build_market_prices(searched, catalog_rows)

    updated, premiums = [], []
    for row_num, whiskey, msrp in catalog_rows:
        mkt = market_prices[row_num]
        ws.cell(row=row_num, column=10).value = mkt
        updated.append(whiskey.strip())
        if abs(mkt - msrp) > 2:
            premiums.append((whiskey.strip(), msrp, mkt, mkt - msrp))

    # Step 3: save and report
    wb.save(CATALOG_PATH)
    print(f"\nUpdated {len(updated)} expressions.\n")

    if premiums:
        print("Expressions where market price differs from MSRP by >$2:")
        print(f"  {'Expression':<48} {'MSRP':>7}  {'Mkt Px':>7}  {'Diff':>7}")
        print("  " + "-" * 74)
        for name, msrp, mkt, diff in sorted(premiums, key=lambda x: -x[3]):
            print(f"  {name:<48} ${msrp:>6.2f}  ${mkt:>6.2f}  +${diff:>5.2f}")
    else:
        print("No expressions with market price > $2 above MSRP.")

    print("\nDone. Street RBEU and Street Tier recalculate automatically in Excel.")


if __name__ == "__main__":
    run()
