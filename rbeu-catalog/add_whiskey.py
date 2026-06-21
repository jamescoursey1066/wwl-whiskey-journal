#!/usr/bin/env python3
"""
xRBEU Catalog - Add Whiskey Agent
Adds a new whiskey entry to the xRBEU Catalog spreadsheet.
Usage: python add_whiskey.py "Whiskey Name" <rating_0_100> <msrp> "notes" "bbb"
  bbb: bottle | bar | bust
Example: python add_whiskey.py "Buffalo Trace" 85 29.99 "Classic benchmark bourbon" "bottle"
"""

import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# ── Constants ──────────────────────────────────────────────────────────────────
import math
SPREADSHEET_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "RBEU Catalog.xlsx")
WT_ROW = 30          # Wild Turkey Rare Breed reference row (NEVER change this)
TOP_RATING = 100.0   # Top-tier anchor: a 100 rating is worth $250 (100-point scale)
TOP_PRICE = 250.00

# Cell formatting to match existing spreadsheet style
CURRENCY_FORMAT = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
RBEU_FORMAT = "0.00"
RATING_FORMAT = "0"

# ── Core calculations ───────────────────────────────────────────────────────────
# Exponential xRBEU (variant s153, recalibrated for two-anchor smoothing):
#     fair_price(R) = WT_PRICE * exp(k * (R - WT_RATING))
#     k = ln(TOP_PRICE / WT_PRICE) / (TOP_RATING - WT_RATING)
#     xRBEU(R, P) = fair_price(R) / P
# Anchors: (WT_RATING, WT_PRICE) and (TOP_RATING=10.0, TOP_PRICE=$250).
# Both anchor points map to xRBEU = 1.0. Values > 1.0 = better-than-fair value.
# WT_RATING / WT_PRICE are read live from cells D30 / E30 to stay in sync.
def calculate_rbeu(rating, price, wt_rating, wt_price):
    """Exponential xRBEU anchored at WT row 30 and at (10.0, $150)."""
    k = math.log(TOP_PRICE / wt_price) / (TOP_RATING - wt_rating)
    fair_price = wt_price * math.exp(k * (rating - wt_rating))
    return fair_price / price

def tier_from_rbeu(rbeu):
    """Map xRBEU value to tier letter grade."""
    if rbeu >= 1.75: return "S"
    elif rbeu >= 1.00: return "A"
    elif rbeu >= 0.85: return "B"
    elif rbeu >= 0.70: return "C"
    elif rbeu >= 0.50: return "D"
    else: return "F"

# ── Spreadsheet logic ───────────────────────────────────────────────────────────
def find_insertion_rows(ws):
    """Find the last row with a whiskey name (col C) and the last row with a rank (col B)."""
    last_named_row = 4
    last_ranked_row = 5
    for row_num in range(5, ws.max_row + 1):
        rank_val = ws.cell(row=row_num, column=2).value   # Column B
        name_val = ws.cell(row=row_num, column=3).value   # Column C
        if name_val is not None:
            last_named_row = row_num
        if rank_val is not None:
            last_ranked_row = row_num
    return last_named_row, last_ranked_row

def apply_cell_style(cell, bold=False, h_align=None, number_format=None):
    """Apply consistent Calibri 11pt styling to a cell."""
    cell.font = Font(name="Calibri", size=11, bold=bold)
    cell.alignment = Alignment(
        horizontal=h_align,
        vertical="center"
    )
    if number_format:
        cell.number_format = number_format

VALID_BBB = {"bottle", "bar", "bust"}

def add_whiskey(name, rating, price, notes="", bbb=""):
    """Add a new whiskey entry to the xRBEU Catalog spreadsheet."""

    if not os.path.exists(SPREADSHEET_PATH):
        print(f"ERROR: Spreadsheet not found at: {SPREADSHEET_PATH}")
        sys.exit(1)

    if bbb and bbb.lower() not in VALID_BBB:
        print(f"ERROR: BBB must be one of: bottle, bar, bust (got '{bbb}')")
        sys.exit(1)
    bbb = bbb.lower() if bbb else ""

    wb = load_workbook(SPREADSHEET_PATH)
    ws = wb.active

    # Read live WT anchor from cells D30 / E30 so display calc matches sheet
    wt_rating = float(ws.cell(row=WT_ROW, column=4).value)
    wt_price = float(ws.cell(row=WT_ROW, column=5).value)

    # Determine where to insert
    last_named_row, last_ranked_row = find_insertion_rows(ws)
    new_row = last_named_row + 1

    # Rank formula: reference the last ranked row (never row 30 which is WT)
    rank_ref = last_ranked_row if last_ranked_row != WT_ROW else WT_ROW - 1

    # ── Column B: Rank ─────────────────────────────────────────────────────────
    cell = ws.cell(row=new_row, column=2, value=f"=B{rank_ref}+1")
    apply_cell_style(cell, bold=True, h_align="right")

    # ── Column C: Whiskey Name ─────────────────────────────────────────────────
    cell = ws.cell(row=new_row, column=3, value=name)
    apply_cell_style(cell)

    # ── Column D: t8ke Rating ──────────────────────────────────────────────────
    cell = ws.cell(row=new_row, column=4, value=float(rating))
    apply_cell_style(cell, h_align="center", number_format=RATING_FORMAT)

    # ── Column E: MSRP ────────────────────────────────────────────────────────
    cell = ws.cell(row=new_row, column=5, value=float(price))
    apply_cell_style(cell, number_format=CURRENCY_FORMAT)

    # ── Column F: xRBEU (exponential, anchored at WT row 30 and R=10/$150) ────
    rbeu_formula = (
        f"=($E${WT_ROW}*EXP((LN({TOP_PRICE}/$E${WT_ROW})/({TOP_RATING}-$D${WT_ROW}))"
        f"*(D{new_row}-$D${WT_ROW})))/E{new_row}"
    )
    cell = ws.cell(row=new_row, column=6, value=rbeu_formula)
    apply_cell_style(cell, bold=True, h_align="center", number_format=RBEU_FORMAT)

    # ── Column G: Tier (IF formula) ───────────────────────────────────────────
    tier_formula = (
        f'=IF(F{new_row}="","",IF(F{new_row}>=1.75,"S",'
        f'IF(F{new_row}>=1,"A",IF(F{new_row}>=0.85,"B",'
        f'IF(F{new_row}>=0.7,"C",IF(F{new_row}>=0.5,"D","F"))))))'
    )
    cell = ws.cell(row=new_row, column=7, value=tier_formula)
    apply_cell_style(cell, h_align="center")

    # ── Column H: BBB (Bottle, Bar, or Bust) ──────────────────────────────────
    if bbb:
        cell = ws.cell(row=new_row, column=8, value=bbb.capitalize())
        apply_cell_style(cell, h_align="center")

    # ── Column I: Notes ───────────────────────────────────────────────────────
    if notes:
        cell = ws.cell(row=new_row, column=9, value=notes)
        apply_cell_style(cell)

    wb.save(SPREADSHEET_PATH)

    # Calculate for display (formula will recalculate in Excel)
    rbeu = calculate_rbeu(rating, price, wt_rating, wt_price)
    tier = tier_from_rbeu(rbeu)

    print(f"Successfully added to xRBEU Catalog (row {new_row}):")
    print(f"  Whiskey : {name}")
    print(f"  Rating  : {rating}")
    print(f"  MSRP    : ${price:.2f}")
    print(f"  xRBEU   : {rbeu:.4f}  →  Tier {tier}")
    if notes:
        print(f"  Notes   : {notes}")
    if bbb:
        print(f"  BBB     : {bbb.capitalize()}")
    print(f"  Row     : {new_row}")

    # Auto-refresh the interactive catalog viewer
    try:
        import generate_catalog_viewer
        generate_catalog_viewer.main()
    except Exception as e:
        print(f"  (warning: catalog_viewer.html not refreshed: {e})")

    return {
        "name": name,
        "rating": rating,
        "price": price,
        "notes": notes,
        "bbb": bbb,
        "rbeu": round(rbeu, 4),
        "tier": tier,
        "row": new_row
    }

# ── Entry point ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)

    result = add_whiskey(
        name=sys.argv[1],
        rating=float(sys.argv[2]),
        price=float(sys.argv[3]),
        notes=sys.argv[4] if len(sys.argv) > 4 else "",
        bbb=sys.argv[5] if len(sys.argv) > 5 else ""
    )
